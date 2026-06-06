"""Tạo file mẫu và kiểm thử ingest + API."""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "backend"))

import pandas as pd
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database import Base
from ingest import ingest_file
from parsers import parse_date_shopee, parse_id_string, parse_percent, parse_vn_number

SAMPLES = ROOT / "tests" / "samples"
SAMPLES.mkdir(parents=True, exist_ok=True)
TEST_DB = ROOT / "data" / "test_datahub.db"


def test_normalization():
    """Kiểm tra helper chuẩn hoá số Việt Nam."""
    assert parse_vn_number("3.293.607.217") == 3293607217.0
    assert parse_vn_number("333.834,10") == 333834.10
    assert parse_percent("7,34%") == 7.34
    assert parse_id_string("7123456789012345678.0") == "7123456789012345678"
    assert parse_date_shopee("01-05-2026-31-05-2026") is None
    assert str(parse_date_shopee("15-05-2026")) == "2026-05-15"
    print("✓ Normalization helpers OK")


def create_tiktok_order_csv():
    path = SAMPLES / "tiktok_orders_sample.csv"
    rows = [
        {
            "Order ID": "7123456789012345678",
            "Order Status": "Đã giao",
            "Order Substatus": "",
            "Cancelation/Return Type": "",
            "SKU ID": "9876543210",
            "Seller SKU": "SKU-001",
            "Product Name": "Sản phẩm A",
            "Variation": "Size M",
            "Quantity": "2",
            "Sku Quantity of return": "0",
            "SKU Unit Original Price": "150.000",
            "SKU Subtotal Before Discount": "300.000",
            "SKU Platform Discount": "30.000",
            "SKU Seller Discount": "0",
            "SKU Subtotal After Discount": "270.000",
            "Shipping Fee After Discount": "15.000",
            "Original Shipping Fee": "30.000",
            "Taxes": "0",
            "Order Amount": "285.000",
            "Order Refund Amount": "0",
            "Created Time": "01/05/2026 10:30:00",
            "Paid Time": "01/05/2026 10:31:00",
            "Shipped Time": "02/05/2026 08:00:00",
            "Delivered Time": "05/05/2026 14:00:00",
            "Cancelled Time": "",
            "Cancel Reason": "",
            "Fulfillment Type": "FBT",
            "Warehouse Name": "WH1",
            "Delivery Option": "Standard",
            "Shipping Provider Name": "J&T",
            "Buyer Username": "buyer1",
            "Province": "Hà Nội",
            "District": "Cầu Giấy",
            "Payment Method": "COD",
            "Weight(kg)": "0.5",
            "Product Category": "Thực phẩm",
        },
        {
            "Order ID": "7123456789012345678",
            "Order Status": "Đã giao",
            "Seller SKU": "SKU-002",
            "Product Name": "Sản phẩm B",
            "Quantity": "1",
            "SKU Subtotal After Discount": "120.000",
            "Created Time": "01/05/2026 10:30:00",
        },
        {
            "Order ID": "7987654321098765432",
            "Order Status": "Đã hủy",
            "Seller SKU": "SKU-003",
            "Product Name": "Sản phẩm C",
            "Quantity": "1",
            "SKU Subtotal After Discount": "50.000",
            "Created Time": "02/05/2026 11:00:00",
        },
    ]
    df = pd.DataFrame(rows)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    print(f"✓ Created {path.name}")
    return path


def create_tiktok_ad_xlsx():
    path = SAMPLES / "Creative_data_20260501-20260531.xlsx"
    df = pd.DataFrame([
        {
            "ID video": "7123456789012345678",
            "Tiêu đề video": "Video quảng cáo 1",
            "Tài khoản TikTok": "@shop",
            "Loại nội dung sáng tạo": "Video",
            "Nguồn video": "Tự tạo",
            "Trạng thái": "Active",
            "Thời gian đăng": "01/05/2026 09:00:00",
            "Chi phí": "1.500.000",
            "Số lượng đơn hàng SKU": "50",
            "Chi phí cho mỗi đơn hàng": "30.000",
            "Doanh thu gộp": "5.000.000",
            "ROI": "3,33",
            "Số lượt hiển thị quảng cáo sản phẩm": "100.000",
            "Số lượt nhấp vào quảng cáo sản phẩm": "2.000",
            "Tỷ lệ nhấp vào quảng cáo sản phẩm": "2,00%",
            "Tỷ lệ chuyển đổi quảng cáo": "2,50%",
            "Tỷ lệ xem video quảng cáo trong 2 giây": "45%",
            "Tỷ lệ xem video quảng cáo trong 6 giây": "30%",
            "Tỷ lệ xem 25% thời lượng video quảng cáo": "25%",
            "Tỷ lệ xem 50% thời lượng video quảng cáo": "15%",
            "Tỷ lệ xem 75% thời lượng video quảng cáo": "10%",
            "Tỷ lệ xem 100% thời lượng video quảng cáo": "5%",
            "Đơn vị tiền tệ": "VND",
        },
    ])
    df.to_excel(path, index=False, engine="openpyxl")
    print(f"✓ Created {path.name}")
    return path


def create_affiliate_creator_xlsx():
    path = SAMPLES / "Creator_List_20260501-20260531.xlsx"
    df = pd.DataFrame([
        {
            "Tên người dùng của nhà sáng tạo": "creator_a",
            "GMV liên kết": "10.000.000",
            "GMV LIVE của liên kết": "3.000.000",
            "GMV video link bán hàng của liên kết": "7.000.000",
            "GMV thẻ sản phẩm của liên kết": "0",
            "Sản phẩm liên kết đã bán": "100",
            "Số món bán ra": "150",
            "Hoa hồng ước tính": "500.000",
            "Phí cố định ước tính": "0",
            "Giá trị đơn hàng trung bình": "100.000",
            "Trang trưng bày sản phẩm liên kết": "5",
            "Đơn hàng liên kết": "100",
            "Tỷ lệ nhấp (CTR)": "3,5%",
            "Lượt hiển thị sản phẩm": "50.000",
            "Khách hàng liên kết trung bình": "80",
            "Buổi LIVE liên kết": "10",
            "Video link bán hàng của liên kết": "20",
            "GMV cộng tác mục tiêu/mở rộng của liên kết": "0",
            "GMV đã hoàn tiền từ liên kết": "200.000",
            "Mặt hàng từ liên kết đã hoàn tiền": "2",
            "Người theo dõi của liên kết": "10.000",
        },
        {
            "Tên người dùng của nhà sáng tạo": "creator_b",
            "GMV liên kết": "5.000.000",
            "Đơn hàng liên kết": "50",
        },
    ])
    df.to_excel(path, index=False, engine="openpyxl")
    print(f"✓ Created {path.name}")
    return path


def create_affiliate_video_xlsx():
    path = SAMPLES / "Creator_Video_List_20260501-20260531.xlsx"
    df = pd.DataFrame([
        {
            "Tên video": "Review sản phẩm A",
            "Ngày đăng video": "01-05-2026",
            "Tên người dùng của nhà sáng tạo": "creator_a",
            "Bình luận của video link bán hàng": "100",
            "Lượt thích của video link bán hàng": "5000",
            "Đơn hàng liên kết": "30",
            "Số món bán ra qua liên kết ": "45",
            "Giá trị đơn hàng trung bình của video link bán hàng": "120.000",
            "GMV": "5.400.000",
            "GMV video link bán hàng của liên kết": "5.400.000",
            "Hoa hồng ước tính": "270.000",
            "Phí cố định ước tính": "0",
            "Khách hàng liên kết trung bình": "25",
            "Mặt hàng từ liên kết đã hoàn tiền": "1",
            "GMV đã hoàn tiền từ liên kết": "120.000",
            "Lượt hiển thị của video link bán hàng": "20.000",
            "CTR của liên kết": "2,5%",
            "GPM của video link bán hàng": "270",
        },
    ])
    df.to_excel(path, index=False, engine="openpyxl")
    print(f"✓ Created {path.name}")
    return path


def create_shopee_xlsx():
    path = SAMPLES / "shopee-shop-stats-20260501-20260531.xlsx"
    wb = Workbook()

    daily_cols = [
        "Ngày", "Tổng doanh số (VND)", "Doanh số không bao gồm trợ giá bởi Shopee",
        "Tổng số đơn hàng", "Doanh số trên mỗi đơn hàng", "Lượt nhấp vào sản phẩm",
        "Số lượt truy cập", "Tỷ lệ chuyển đổi đơn hàng", "Đơn đã hủy", "Doanh số đơn hủy",
        "Đơn đã hoàn trả / hoàn tiền", "Doanh số các đơn Trả hàng/Hoàn tiền",
        "số người mua", "số người mua mới", "số người mua hiện tại",
        "số người mua tiềm năng", "Tỉ lệ quay lại của người mua",
    ]

    for sheet_name in ("Đơn hàng đã đặt", "Đơn đã xác nhận", "Đơn Đã Thanh Toán"):
        ws = wb.create_sheet(sheet_name)
        ws.append(daily_cols)
        ws.append(["01-05-2026-31-05-2026", "50.000.000"] + [""] * (len(daily_cols) - 2))
        ws.append([""] * len(daily_cols))
        ws.append(daily_cols)
        for day in ("01-05-2026", "02-05-2026", "03-05-2026"):
            ws.append([
                day, "1.500.000", "1.400.000", "10", "150.000",
                "500", "200", "5,00%", "1", "100.000",
                "0", "0", "8", "3", "5", "2", "60%",
            ])

    ws_prod = wb.create_sheet("Theo sản phẩm (placed)")
    ws_prod.append(["Báo cáo sản phẩm"])
    ws_prod.append([
        "Mã sản phẩm", "Sản phẩm", "Tình trạng", "Tỷ lệ doanh số",
        "Doanh số (VND)", "Lượt hiển thị", "Lượt nhấp", "Tổng số đơn hàng",
        "Sản phẩm", "CTR", "Tỷ lệ chuyển đổi", "Doanh số/đơn", "Người mua",
    ])
    ws_prod.append(["12345678901", "SP A", "Normal", "60%", "900.000", "1000", "100", "6", "8", "10%", "6%", "150.000", "5"])
    ws_prod.append(["98765432109", "SP B", "Normal", "40%", "600.000", "800", "80", "4", "5", "10%", "5%", "150.000", "3"])

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(path)
    print(f"✓ Created {path.name}")
    return path


def run_ingest_tests():
    """Nạp tất cả file mẫu vào DB test."""
    if TEST_DB.exists():
        TEST_DB.unlink()

    engine = create_engine(f"sqlite:///{TEST_DB}")
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    db = Session()

    files = [
        create_tiktok_order_csv(),
        create_tiktok_ad_xlsx(),
        create_affiliate_creator_xlsx(),
        create_affiliate_video_xlsx(),
        create_shopee_xlsx(),
    ]

    for fp in files:
        content = fp.read_bytes()
        result = ingest_file(db, fp.name, content)
        assert result["status"] == "success", f"Failed: {fp.name} -> {result}"
        print(f"✓ Ingested {fp.name}: {result['row_count']} rows")

    dup = ingest_file(db, files[0].name, files[0].read_bytes())
    assert dup["status"] == "duplicate", f"Expected duplicate, got {dup}"
    print("✓ Duplicate detection OK")

    db.close()
    print(f"\n✓ All ingest tests passed. DB: {TEST_DB}")


if __name__ == "__main__":
    test_normalization()
    run_ingest_tests()
